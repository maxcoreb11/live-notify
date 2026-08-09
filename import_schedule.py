"""
Excel -> schedule.json Converter
=================================
แปลงตารางถ่ายทอดสดจากไฟล์ Excel (.xlsx) ให้เป็น schedule.json ที่ notify_bot.py ใช้

วิธีใช้ (รันในเครื่องตัวเอง ไม่ได้รันบน GitHub Actions):
    pip install openpyxl
    python import_schedule.py "Stream Schedule.xlsx"
    python import_schedule.py "Stream Schedule.xlsx" --sheet "กันยายน 2569"
    python import_schedule.py "Stream Schedule.xlsx" --append   # เพิ่มต่อท้ายของเดิม ไม่ล้างทิ้ง

ตัวสคริปต์จะ:
    - หาแถวหัวตารางเอง (แถวที่มีคำว่า "วันที่") แล้วแมปคอลัมน์ตามชื่อหัวตาราง
    - แตกช่องเวลาที่มีหลายรอบ เช่น "18:00 / 20:30" ออกเป็นหลายรายการ
    - รายการที่เวลายังไม่แน่นอน (เช่น "รอประกาศ") ถูกแยกไป needs_review.json แทน
    - รักษาสถานะ notified_leads ของรายการเดิมที่ตรงกันไว้ (จะได้ไม่แจ้งเตือนซ้ำ)
"""

import argparse
import json
import os
import re
import sys
from datetime import date as date_cls, datetime, time as time_cls

try:
    import openpyxl
except ImportError:
    sys.exit("ต้องติดตั้ง openpyxl ก่อน:  pip install openpyxl")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCHEDULE_PATH = os.path.join(BASE_DIR, "schedule.json")
REVIEW_PATH = os.path.join(BASE_DIR, "needs_review.json")

DAY_TH = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]

# ชื่อหัวตารางที่ยอมรับ -> ชื่อฟิลด์ภายใน
HEADER_MAP = {
    "วันที่": "date",
    "วัน": "day_th",
    "เวลา (ไทย)": "time",
    "เวลา": "time",
    "ประเภท": "category",
    "รายการ": "title",
    "คู่ / รายละเอียด": "detail",
    "คู่/รายละเอียด": "detail",
    "ดูที่ไหน": "channel",
    "ฟรีทีวี?": "free_tv",
    "ฟรีทีวี": "free_tv",
    "P'เปรี้ยว": "club",
}

TIME_RE = re.compile(r"^([0-2]?\d)[:.]([0-5]\d)$")


def cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value):
    """คืนค่า datetime.date หรือ None ถ้าแปลงไม่ได้"""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_cls):
        return value
    text = cell_str(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def parse_times(value):
    """แปลงช่องเวลาเป็น list ของ 'HH:MM'

    รองรับ: datetime.time, '19:00', '03:00:00', '18:00 / 20:30',
            '18:30 / 20:30 (ฟรี Ch7 20:30)', '18:30-20:30'
    คืนค่า [] ถ้าไม่มีเวลาที่ใช้ได้ (เช่น 'รอประกาศ')
    """
    if isinstance(value, time_cls):
        return [f"{value.hour:02d}:{value.minute:02d}"]
    if isinstance(value, datetime):
        return [f"{value.hour:02d}:{value.minute:02d}"]

    text = cell_str(value)
    if not text:
        return []

    # ตัดหมายเหตุในวงเล็บออกก่อน เช่น "(ฟรี Ch7 20:30)" ไม่ใช่รอบถ่ายทอดใหม่
    text = re.sub(r"\([^)]*\)", " ", text)

    times, seen = [], set()
    for chunk in re.split(r"[/,\-–]| และ ", text):
        chunk = chunk.strip()
        if not chunk:
            continue
        # ตัดวินาทีทิ้ง เช่น '03:00:00' -> '03:00'
        chunk = re.sub(r"^(\d{1,2}:\d{2}):\d{2}$", r"\1", chunk)
        m = TIME_RE.match(chunk)
        if not m:
            continue
        hh, mm = int(m.group(1)), int(m.group(2))
        if hh > 23:
            continue
        formatted = f"{hh:02d}:{mm:02d}"
        if formatted not in seen:
            seen.add(formatted)
            times.append(formatted)
    return times


def find_header_row(ws, max_scan=15):
    for row in ws.iter_rows(min_row=1, max_row=max_scan):
        values = [cell_str(c.value) for c in row]
        if "วันที่" in values and any(v in ("ประเภท", "รายการ") for v in values):
            return row[0].row, {
                HEADER_MAP[v]: i for i, v in enumerate(values) if v in HEADER_MAP
            }
    return None, None


def pick_sheet(wb, requested):
    if requested:
        if requested not in wb.sheetnames:
            sys.exit(f"ไม่พบชีตชื่อ '{requested}' — ชีตที่มี: {wb.sheetnames}")
        return wb[requested]
    # เลือกชีตแรกที่มีหัวตารางถูกต้อง (ชีตเดือนล่าสุดมักอยู่ซ้ายสุด)
    for ws in wb.worksheets:
        if find_header_row(ws)[0]:
            return ws
    sys.exit("ไม่พบชีตที่มีหัวตาราง 'วันที่' เลย")


def event_key(event):
    return (event["date"], event["time"], event["title"], event["detail"])


def load_existing_leads():
    """เก็บสถานะ notified_leads เดิมไว้ กันแจ้งเตือนซ้ำเมื่อ import ทับ"""
    if not os.path.exists(SCHEDULE_PATH):
        return {}
    try:
        with open(SCHEDULE_PATH, "r", encoding="utf-8") as f:
            old = json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}
    return {event_key(e): e.get("notified_leads", []) for e in old if "date" in e}


def convert(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = pick_sheet(wb, sheet_name)
    header_row, cols = find_header_row(ws)
    if not header_row:
        sys.exit(f"ชีต '{ws.title}' ไม่มีแถวหัวตาราง 'วันที่'")

    missing = {"date", "time", "category", "title"} - set(cols)
    if missing:
        sys.exit(f"ชีต '{ws.title}' ขาดคอลัมน์: {', '.join(sorted(missing))}")

    events, review = [], []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        def col(name):
            idx = cols.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        event_date = parse_date(col("date"))
        if event_date is None:
            continue  # แถวว่างหรือแถวหมายเหตุท้ายตาราง

        title = cell_str(col("title"))
        detail = cell_str(col("detail"))
        if not title and not detail:
            continue

        base = {
            "date": event_date.isoformat(),
            "day_th": DAY_TH[event_date.weekday()],
            "category": cell_str(col("category")),
            "title": title,
            "detail": detail,
            "channel": cell_str(col("channel")),
            "free_tv": cell_str(col("free_tv")),
            "club": cell_str(col("club")),
        }

        times = parse_times(col("time"))
        if not times:
            review.append({**base, "raw_time": cell_str(col("time")),
                           "reason": "ไม่มีเวลาที่แน่นอน - ต้องกรอกเวลาเองใน schedule.json"})
            continue

        for t in times:
            events.append({**base, "time": t, "notified_leads": []})

    events.sort(key=lambda e: (e["date"], e["time"], e["title"]))
    return ws.title, events, review


def main():
    ap = argparse.ArgumentParser(description="แปลงตาราง Excel เป็น schedule.json")
    ap.add_argument("xlsx", help="ไฟล์ .xlsx ที่มีตารางถ่ายทอดสด")
    ap.add_argument("--sheet", help="ชื่อชีต เช่น 'สิงหาคม 2569' (ไม่ระบุ = ชีตแรกที่ใช้ได้)")
    ap.add_argument("--append", action="store_true",
                    help="เพิ่มต่อท้าย schedule.json เดิม แทนการเขียนทับทั้งไฟล์")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit(f"ไม่พบไฟล์ {args.xlsx}")

    sheet_title, events, review = convert(args.xlsx, args.sheet)

    old_leads = load_existing_leads()
    for e in events:
        e["notified_leads"] = old_leads.get(event_key(e), [])

    if args.append:
        with open(SCHEDULE_PATH, "r", encoding="utf-8") as f:
            existing = json.load(f)
        seen = {event_key(e) for e in events}
        events = [e for e in existing if event_key(e) not in seen] + events
        events.sort(key=lambda e: (e["date"], e["time"], e["title"]))

    with open(SCHEDULE_PATH, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=2)
        f.write("\n")
    with open(REVIEW_PATH, "w", encoding="utf-8") as f:
        json.dump(review, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"ชีตที่ใช้: {sheet_title}")
    print(f"schedule.json     : {len(events)} รายการ")
    print(f"needs_review.json : {len(review)} รายการ (เวลายังไม่แน่นอน)")
    kept = sum(1 for e in events if e["notified_leads"])
    if kept:
        print(f"คงสถานะแจ้งเตือนเดิมไว้ {kept} รายการ")


if __name__ == "__main__":
    main()
